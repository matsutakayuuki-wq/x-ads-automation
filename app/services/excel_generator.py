"""X Ads Editor 形式の Excel ファイル生成"""
from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Project, SubmissionBatch, SubmissionCampaign

# X Ads Editor のキャンペーンシート 102カラム（日本語ヘッダー）
EXCEL_COLUMNS = [
    "キャンペーンID",                          # 0  (A)
    "お支払い方法ID",                          # 1  (B)
    "キャンペーン名",                          # 2  (C)
    "キャンペーン開始日",                       # 3  (D)
    "キャンペーン終了日",                       # 4  (E)
    "広告キャンペーン予算の最適化",               # 5  (F)
    "キャンペーンステータス",                    # 6  (G)
    "配信",                                    # 7  (H)
    "キャンペーン総予算",                       # 8  (I)
    "広告キャンペーンの日別予算",                # 9  (J)
    "広告代理店クレジットライン発注番号",          # 10 (K)
    "キャンペーンのフリークエンシー上限",          # 11 (L)
    "キャンペーンのフリークエンシー上限が適用される期間",  # 12 (M)
    "広告グループID",                          # 13 (N)
    "キャンペーンの目的",                       # 14 (O)
    "広告グループ名",                          # 15 (P)
    "広告グループの開始時刻",                    # 16 (Q)
    "広告グループの終了時刻",                    # 17 (R)
    "広告グループの状態",                       # 18 (S)
    "広告グループの総予算",                      # 19 (T)
    "広告グループの日別予算",                    # 20 (U)
    "配信",                                    # 21 (V)
    "広告グループのフリークエンシー上限",          # 22 (W)
    "広告グループのフリークエンシー上限が適用される期間",  # 23 (X)
    "目標",                                    # 24 (Y)
    "広告グループの配信先",                      # 25 (Z)
    "プロモ商品タイプ",                         # 26 (AA)
    "ウェブサイトコンバージョンタグID",            # 27 (AB)
    "お支払い方法",                             # 28 (AC)
    "広告主ドメイン",                           # 29 (AD)
    "入札戦略",                                # 30 (AE)
    "入札額",                                  # 31 (AF)
    "情報開示のタイプ",                         # 32 (AG)
    "情報開示のテキスト",                       # 33 (AH)
    "IABカテゴリー",                            # 34 (AI)
    "アプリ",                                  # 35 (AJ)
    "Amplifyプログラム",                        # 36 (AK)
    "性別",                                    # 37 (AL)
    "年齢",                                    # 38 (AM)
    "場所",                                    # 39 (AN)
    "フォロワーターゲティング",                  # 40 (AO)
    "フォロワーターゲティングの類似ターゲティング",  # 41 (AP)
    "プラットフォーム",                         # 42 (AQ)
    "ユーザーのOSバージョン",                    # 43 (AR)
    "ユーザーの端末",                           # 44 (AS)
    "Wi-Fiのみ",                               # 45 (AT)
    "言語",                                    # 46 (AU)
    "ユーザーの興味関心",                       # 47 (AV)
    "携帯電話会社",                             # 48 (AW)
    "端末アクティベーション期間",                # 49 (AX)
    "完全一致キーワード",                       # 50 (AY)
    "除外する完全一致キーワード",                # 51 (AZ)
    "部分一致キーワード",                       # 52 (BA)
    "順不同キーワード",                         # 53 (BB)
    "除外する順不同キーワード",                  # 54 (BC)
    "フレーズキーワード",                       # 55 (BD)
    "除外するフレーズキーワード",                # 56 (BE)
    "テイラードオーディエンスリスト",             # 57 (BF)
    "除外するテイラードオーディエンスリスト",      # 58 (BG)
    "テイラードオーディエンスの類似オーディエンス",  # 59 (BH)
    "テイラードオーディエンス (モバイルアプリから)",  # 60 (BI)
    "除外するテイラードオーディエンス (モバイルアプリから)",  # 61 (BJ)
    "テイラードオーディエンス (モバイルアプリから) の類似オーディエンス",  # 62 (BK)
    "テイラードオーディエンスのウェブサイト訪問者",  # 63 (BL)
    "除外するテイラードオーディエンスのウェブサイト訪問者",  # 64 (BM)
    "テイラードオーディエンスのウェブサイト訪問者の類似オーディエンス",  # 65 (BN)
    "インストールされているアプリのカテゴリー",    # 66 (BO)
    "インストールされているアプリのカテゴリーの類似カテゴリー",  # 67 (BP)
    "TAP - 除外するアプリ",                     # 68 (BQ)
    "予約済み",                                # 69 (BR)
    "TAP - パブリッシャーアプリのカテゴリー",      # 70 (BS)
    "TV番組",                                  # 71 (BT)
    "イベントターゲティングID",                  # 72 (BU)
    "キャンペーンのリターゲティング",             # 73 (BV)
    "オーガニックなツイートのリターゲティング",     # 74 (BW)
    "エンゲージメントタイプをリターゲティング",     # 75 (BX)
    "ツイートID",                              # 76 (BY)
    "予約投稿ツイートID",                       # 77 (BZ)
    "プロモアカウントID",                       # 78 (CA)
    "予約済み",                                # 79 (CB)
    "TAPメディアクリエイティブアプリID",           # 80 (CC)
    "TAPメディアクリエイティブランディングURL",     # 81 (CD)
    "メディアクリエイティブID",                  # 82 (CE)
    "予約済み",                                # 83 (CF)
    "予約済み",                                # 84 (CG)
    "Google Campaign Manager tags",            # 85 (CH)
    "予約済み (2)",                             # 86 (CI)
    "予約済み（3）",                            # 87 (CJ)
    "フレキシブルオーディエンス",                 # 88 (CK)
    "柔軟なオーディエンスを除外",                # 89 (CL)
    "柔軟なオーディエンスの類似オーディエンス",     # 90 (CM)
    "Amplifyプログラムから自動プロモーション",     # 91 (CN)
    "予約済み",                                # 92 (CO)
    "予約済み",                                # 93 (CP)
    "類似オーディエンスの拡張設定",              # 94 (CQ)
    "会話",                                    # 95 (CR)
    "ターゲティングするパブリッシャーアカウント",   # 96 (CS)
    "除外したパブリッシャーアカウント",            # 97 (CT)
    "除外したIABカテゴリー",                     # 98 (CU)
    "標準カテゴリー",                           # 99 (CV)
    "Amplifyプレロールプレミアムカテゴリー",       # 100 (CW)
    "予約済み",                                # 101 (CX)
]

# X Ads Editor の Objective マッピング
# API の WEBSITE_CONVERSIONS は Excel では WEBSITE_CLICKS を使用
OBJECTIVE_MAP_EXCEL = {
    "WEBSITE_CLICKS": "WEBSITE_CLICKS",
    "WEBSITE_CONVERSIONS": "WEBSITE_CLICKS",  # API v12 と同じく WEBSITE_CLICKS を使用
    "APP_INSTALLS": "APP_INSTALLS",
    "ENGAGEMENTS": "ENGAGEMENTS",
    "REACH": "REACH",
    "VIDEO_VIEWS": "VIDEO_VIEWS",
    "FOLLOWERS": "FOLLOWERS",
    "APP_ENGAGEMENTS": "APP_ENGAGEMENTS",
}

# 入札戦略: UI/API形式 → Excel形式
BID_STRATEGY_MAP_EXCEL = {
    "AUTO": "AUTO_BID",
    "TARGET": "AUTO_BID_WITH_TARGET",
    "MAX": "MAX_BID",
    # すでにExcel形式の場合
    "AUTO_BID": "AUTO_BID",
    "MAX_BID": "MAX_BID",
    "AUTO_BID_WITH_TARGET": "AUTO_BID_WITH_TARGET",
    "NO_BID": "NO_BID",
}

# Placement: API形式 → Excel形式（セミコロン区切り）
PLACEMENT_MAP_EXCEL = {
    "ALL_ON_TWITTER": "ALL_ON_TWITTER",
    "TWITTER_TIMELINE": "TIMELINES",
    "TWITTER_SEARCH": "SEARCH_RESULTS",
    "TWITTER_PROFILE": "PROFILES",
    "TWITTER_REPLIES": "REPLIES",
    "TWITTER_MEDIA_VIEWER": "MEDIA_VIEWER",
    "PUBLISHER_NETWORK": "PUBLISHER_NETWORK",
    # すでにExcel形式の場合
    "TIMELINES": "TIMELINES",
    "SEARCH_RESULTS": "SEARCH_RESULTS",
    "PROFILES": "PROFILES",
    "REPLIES": "REPLIES",
    "MEDIA_VIEWER": "MEDIA_VIEWER",
}

# X Ads API targeting_value (hex) → Ads Editor ID マッピング
# syntax_sheet_export.xlsx の (3) Locations シートから取得した公式フォーマット
LOCATION_HEX_TO_EDITOR: dict[str, str] = {
    # Country
    "06ef846bfc783874": "i35:Japan",
    # 47 Prefectures (Regions) - Ads Editor syntax sheet の Ads Editor ID 列と完全一致
    "0b89db31d164a17d": "i30731:北海道 / Hokkaido, JP",
    "1de05e90db6fde15": "i30734:青森県 / Aomori-ken, JP",
    "516fad81ed9abcc2": "i30741:岩手県 / Iwate-ken, JP",
    "1cff59592a4767e9": "i30732:宮城県 / Miyagi-ken, JP",
    "975c45ff265eb77c": "i30757:秋田県 / Akita-ken, JP",
    "4a9a5111024ffa58": "i30738:山形県 / Yamagata-ken, JP",
    "5e921369a11e38d5": "i30744:福島県 / Fukushima-ken, JP",
    "cb7c6e9092251aa1": "i30766:茨城県 / Ibaraki-ken, JP",
    "9452db4fb01f0432": "i30756:栃木県 / Tochigi-ken, JP",
    "00a8aa111d38316c": "i30728:群馬県 / Gunma-ken, JP",
    "6eb3dcfadbbe0c68": "i30750:埼玉県 / Saitama-ken, JP",
    "7562529145a9ed1f": "i30751:千葉県 / Chiba-ken, JP",
    "a56612250c754f23": "i30761:東京都 / Tokyo-to, JP",
    "5f3279ed753778b7": "i30745:神奈川県 / Kanagawa-ken, JP",
    "5af9c3e8dadd043d": "i30743:新潟県 / Niigata-ken, JP",
    "c005c6ef5d97c9da": "i30764:富山県 / Toyama-ken, JP",
    "1d059cea3e433d3d": "i30733:石川県 / Ishikawa-ken, JP",
    "d3bdee61e7cfba0c": "i30767:福井県 / Fukui-ken, JP",
    "a20dcb31ad69d661": "i30759:山梨県 / Yamanashi-ken, JP",
    "f28ae4f6babdb2b5": "i30771:長野県 / Nagano-ken, JP",
    "a3e6429d33900d31": "i30760:岐阜県 / Gifu-ken, JP",
    "d80142cda25d6767": "i30769:静岡県 / Shizuoka-ken, JP",
    "c68b1ffd6bd34468": "i30765:愛知県 / Aichi-ken, JP",
    "f9170e3707e30162": "i30772:三重県 / Mie-ken, JP",
    "287821abb712dd3b": "i30735:滋賀県 / Shiga-ken, JP",
    "d4255b2b43cbf2cc": "i30768:京都 / 京都府 / Kyoto-fu, JP",
    "84316acd652607fa": "i30754:大阪府 / Osaka-fu, JP",
    "46cd5ede80186a9c": "i30737:兵庫県 / Hyogo-ken, JP",
    "6836153322ac8f20": "i30747:奈良県 / Nara-ken, JP",
    "631253d45931eb36": "i30746:和歌山県 / Wakayama-ken, JP",
    "4bb7c88397417f82": "i30739:鳥取県 / Tottori-ken, JP",
    "fd9c584b35c83605": "i30774:島根県 / Shimane-ken, JP",
    "b9f3fc68dd8f717b": "i30763:岡山県 / Okayama-ken, JP",
    "39834ee320359393": "i30736:広島県 / Hiroshima-ken, JP",
    "ab43b3b8a7593bb0": "i30762:山口県 / Yamaguchi-ken, JP",
    "4efce255445dc26a": "i30740:徳島県 / Tokushima-ken, JP",
    "9411fa3e127a9e37": "i30755:香川県 / Kagawa-ken, JP",
    "df55059f8045566b": "i30770:愛媛県 / Ehime-ken, JP",
    "00af8e922b6236ea": "i30729:高知県 / Kochi-ken, JP",
    "684cc1cfd89cacef": "i30748:福岡県 / Fukuoka-ken, JP",
    "6b52871b45b5b261": "i30749:佐賀県 / Saga-ken, JP",
    "59856611bf9bfb97": "i30742:長崎県 / Nagasaki-ken, JP",
    "a160a0ba64b9b2b8": "i30758:熊本県 / Kumamoto-ken, JP",
    "84166be9996a2df5": "i30753:大分県 / Oita-ken, JP",
    "82564af6cbb58e75": "i30752:宮崎県 / Miyazaki-ken, JP",
    "fc60aa4eb7499eb1": "i30773:鹿児島県 / Kagoshima-ken, JP",
    "052e049119fd8da1": "i30730:沖縄県 / Okinawa-ken, JP",
    # 地方（Regions - Area レベル）
    "d473ed704dbcd4a5": "i461:関東 / Kanto Area, JP",
    "d79b76231f5f6ffa": "i463:中部 / Chubu Area, JP",
    "d319ca572780f7e0": "i517:近畿 / Kinki Area, JP",
}

# X Ads Editor で有効な年齢レンジ
VALID_AGE_RANGES = {
    "AGE_OVER_13", "AGE_OVER_18", "AGE_OVER_20", "AGE_OVER_21",
    "AGE_OVER_25", "AGE_OVER_30", "AGE_OVER_50",
    "AGE_13_TO_19", "AGE_13_TO_24", "AGE_13_TO_29", "AGE_13_TO_34",
    "AGE_13_TO_39", "AGE_13_TO_49", "AGE_13_TO_54",
    "AGE_18_TO_24", "AGE_18_TO_34", "AGE_18_TO_49", "AGE_18_TO_54",
    "AGE_20_TO_29", "AGE_20_TO_34", "AGE_20_TO_39", "AGE_20_TO_49",
    "AGE_21_TO_34", "AGE_21_TO_49", "AGE_21_TO_54",
    "AGE_25_TO_49", "AGE_25_TO_54",
    "AGE_30_TO_39", "AGE_30_TO_49",
    "AGE_35_TO_49", "AGE_35_TO_54",
    "AGE_40_TO_49",
}

# 無効な年齢レンジ → 最も近い有効レンジへのマッピング
AGE_RANGE_FALLBACK = {
    "AGE_20_TO_54": "AGE_20_TO_49",
}


def _api_id_to_editor(value: Optional[str]) -> Optional[str]:
    """X Ads API の base36 ID を Editor 形式（i + 10進数）に変換。

    例: '1a1hvg' → 'i77332156', 'zzon8' → 'i60451460'
    すでに 'i' + 数字 形式の場合はそのまま返す。
    """
    if not value:
        return value
    # すでに Editor 形式なら変換不要
    if value.startswith("i") and value[1:].isdigit():
        return value
    try:
        decimal_id = int(value, 36)
        return f"i{decimal_id}"
    except (ValueError, TypeError):
        return value  # 変換できない場合はそのまま


def _tweet_ids_to_editor(value: Optional[str]) -> Optional[str]:
    """ツイートIDをEditor形式に変換（i + 数値ID、セミコロン区切り）"""
    if not value:
        return None
    # JSON配列を試す
    try:
        items = json.loads(value)
        if isinstance(items, list):
            return ";".join(f"i{tid}" if not str(tid).startswith("i") else str(tid) for tid in items)
    except (json.JSONDecodeError, TypeError):
        pass
    # セミコロン or カンマ区切り
    if ";" in value:
        items = [t.strip() for t in value.split(";") if t.strip()]
    else:
        items = [t.strip() for t in value.split(",") if t.strip()]
    return ";".join(f"i{tid}" if not tid.startswith("i") else tid for tid in items)


def _json_to_semicolon(value: Optional[str]) -> str:
    """JSON配列をセミコロン区切りの文字列に変換"""
    if not value:
        return ""
    try:
        items = json.loads(value)
        if isinstance(items, list):
            return ";".join(str(i) for i in items)
        return str(items)
    except (json.JSONDecodeError, TypeError):
        return str(value)


def _locations_to_excel(value: Optional[str]) -> str:
    """地域ターゲティングをAds Editor形式に変換。

    新形式（JSON hex配列）: ["06ef846bfc783874","5f3279ed753778b7",...]
      → hex IDをEditor形式 (i30745:神奈川県 / Kanagawa-ken, JP) に変換してセミコロン区切り
    旧形式（i+数字:名前）: i30745:神奈川県 / Kanagawa-ken, JP;...
      → そのまま返す（既にEditor互換フォーマット）
    """
    if not value:
        return ""

    # JSON配列（新形式）
    try:
        items = json.loads(value)
        if isinstance(items, list):
            editor_ids = []
            for hex_id in items:
                hex_str = str(hex_id).strip()
                editor_id = LOCATION_HEX_TO_EDITOR.get(hex_str)
                if editor_id:
                    editor_ids.append(editor_id)
                # マッピングにないhex IDはスキップ（不明な値を入れるとEditorエラー）
            return ";".join(editor_ids)
    except (json.JSONDecodeError, TypeError):
        pass

    # 旧形式（セミコロン区切り）: そのまま返す（i+数字:名前 形式）
    return str(value)


def _age_range_to_excel(value: Optional[str]) -> str:
    """年齢レンジをExcel有効値に変換（セミコロン区切り）"""
    if not value:
        return ""
    # セミコロン区切り or 単一値
    if ";" in value:
        items = [a.strip() for a in value.split(";") if a.strip()]
    else:
        # JSON配列を試す
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                items = [str(a).strip() for a in parsed]
            else:
                items = [str(parsed).strip()]
        except (json.JSONDecodeError, TypeError):
            items = [value.strip()]

    # 有効値に変換
    result = []
    for age in items:
        if age in VALID_AGE_RANGES:
            result.append(age)
        elif age in AGE_RANGE_FALLBACK:
            result.append(AGE_RANGE_FALLBACK[age])
        # 無効な値はスキップ
    return ";".join(result)


def _placements_to_excel(value: Optional[str], *, has_video: bool = False) -> str:
    """プレースメントをExcel形式に変換（セミコロン区切り）

    MEDIA_VIEWER は動画広告のみ対応のため、has_video=False の場合は除外する。
    """
    if not value:
        return "ALL_ON_TWITTER"

    mapped: list[str] = []

    # JSON配列
    parsed = False
    try:
        items = json.loads(value)
        if isinstance(items, list):
            mapped = [PLACEMENT_MAP_EXCEL.get(p.strip(), p.strip()) for p in items]
            parsed = True
    except (json.JSONDecodeError, TypeError):
        pass

    # セミコロン or カンマ区切り
    if not parsed:
        if ";" in value:
            items = [p.strip() for p in value.split(";") if p.strip()]
        else:
            items = [p.strip() for p in value.split(",") if p.strip()]
        mapped = [PLACEMENT_MAP_EXCEL.get(p, p) for p in items]

    # MEDIA_VIEWER は動画広告のみ有効。画像広告の場合は除外する
    if not has_video:
        mapped = [p for p in mapped if p != "MEDIA_VIEWER"]

    return ";".join(mapped) if mapped else "ALL_ON_TWITTER"


JST = timezone(timedelta(hours=9))


def _format_editor_datetime(value: Optional[str]) -> Optional[str]:
    """日時をX Ads Editor形式 'dd-MMM-yyyy HH:mm' に変換。
    空の場合はデフォルト（現在+5分）を使用。

    例: '2026-03-10T15:00:00+09:00' → '10-Mar-2026 15:00'
    """
    if value:
        # ISO 8601形式やdatetime-local形式をパース
        for fmt in ["%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S+09:00",
                     "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"]:
            try:
                dt = datetime.strptime(value.replace("+09:00", ""), fmt.replace("%z", "").replace("+09:00", ""))
                return dt.strftime("%d-%b-%Y %H:%M")
            except ValueError:
                continue
        # パースできない場合はそのまま返す
        return value
    # 未指定の場合: 現在時刻+5分
    dt = datetime.now(JST) + timedelta(minutes=5)
    return dt.strftime("%d-%b-%Y %H:%M")


def _budget_to_number(value: Optional[int], allow_zero: bool = True) -> Optional[float]:
    """予算値をfloatに変換（Excelの数値セル用）"""
    if value is None:
        return 0.0 if allow_zero else None
    if value == 0:
        return 0.0 if allow_zero else None
    return float(value)


def _gender_to_excel(value: Optional[str]) -> str:
    """性別をExcel形式に変換（元テンプレートに合わせ大文字）"""
    if not value or value.upper() == "ANY":
        return None  # ANYや空は設定しない
    return value.upper()  # MALE / FEMALE


def _platforms_to_excel(value: Optional[str]) -> str:
    """プラットフォームをAds Editor形式に変換（syntax sheet準拠: IOS, ANDROID, DESKTOP）"""
    if not value:
        return ""
    mapping = {
        "0": "IOS",
        "1": "ANDROID",
        "4": "DESKTOP",
        # 既にEditor形式の場合（旧データ互換）
        "iOS": "IOS",
        "IOS": "IOS",
        "Android": "ANDROID",
        "ANDROID": "ANDROID",
        "Desktop": "DESKTOP",
        "DESKTOP": "DESKTOP",
    }
    try:
        items = json.loads(value)
        if isinstance(items, list):
            mapped = [mapping.get(str(i), str(i)) for i in items]
            return ";".join(mapped)
    except (json.JSONDecodeError, TypeError):
        pass
    return str(value)


class ExcelGenerator:
    """X Ads Editor 形式の Excel ファイル生成"""

    def generate(
        self,
        batch: SubmissionBatch,
        project: Optional[Project] = None,
    ) -> BytesIO:
        """SubmissionBatch から Excel ワークブックを生成"""
        wb = Workbook()

        # メインシート: Campaigns
        ws = wb.active
        ws.title = "Campaigns"
        self._build_campaigns_sheet(ws, batch.campaigns, project)

        # 参照シート
        ws2 = wb.create_sheet("お支払い方法ID (編集不可)")
        self._build_payment_sheet(ws2, batch.campaigns, project)

        ws3 = wb.create_sheet("コンバージョンタグID (編集不可)")
        self._build_conversion_tags_sheet(ws3, batch.campaigns)

        ws4 = wb.create_sheet("テイラードオーディエンスID (編集不可)")
        self._build_audiences_sheet(ws4)

        ws5 = wb.create_sheet("メディアクリエイティブID (編集不可)")
        self._build_media_sheet(ws5)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _build_campaigns_sheet(
        self,
        ws,
        campaigns: list[SubmissionCampaign],
        project: Optional[Project],
    ):
        """キャンペーンシートを構築"""
        # ヘッダー行
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(bold=True, size=10)

        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", wrap_text=True)

        # データ行
        for row_idx, campaign in enumerate(campaigns, 2):
            seq = row_idx - 1  # 1, 2, 3, ...
            row_data = self._campaign_to_row(campaign, project, seq)
            for col_idx, value in enumerate(row_data, 1):
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 列幅を調整
        for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(str(EXCEL_COLUMNS[col_idx - 1]))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, min(len(str(cell.value)), 40))
            ws.column_dimensions[letter].width = max_len + 2

        # ヘッダー行を固定
        ws.freeze_panes = "A2"

    def _campaign_to_row(
        self,
        c: SubmissionCampaign,
        project: Optional[Project],
        seq: int = 1,
    ) -> list:
        """SubmissionCampaign を 102要素のリストに変換"""
        row = [None] * len(EXCEL_COLUMNS)

        # === Campaign-level (0-12) ===
        # 新規キャンペーン: [New C1], [New C2], ... の形式
        row[0] = f"New C{seq}"                                        # キャンペーンID (A)
        row[1] = _api_id_to_editor(c.funding_instrument_id)             # お支払い方法ID (B)
        row[2] = c.campaign_name                                      # キャンペーン名 (C)
        start_dt = _format_editor_datetime(c.start_time if c.start_time else None)
        end_dt = _format_editor_datetime(c.end_time) if c.end_time else None
        row[3] = start_dt                                             # キャンペーン開始日 (D)
        row[4] = end_dt                                               # キャンペーン終了日 (E)
        row[5] = "LINE_ITEM"                                             # 予算の最適化 (F) → アドグループ単位で予算管理
        row[6] = "PAUSED"                                             # キャンペーンステータス (G) → 一時停止
        row[7] = True                                                 # 配信 (H) → 標準配信(True) ※False=集中配信はLINE_ITEM時不可
        row[8] = 0.0                                                  # キャンペーン総予算 (I) → 0
        row[9] = 0.0                                                  # キャンペーン日別予算 (J) → 0（予算はアドグループ側）

        # === Ad Group-level (13-36) ===
        # row[13] 広告グループID (N) → 新規は空欄
        objective = OBJECTIVE_MAP_EXCEL.get(c.campaign_objective, c.campaign_objective)
        row[14] = objective                                           # キャンペーンの目的 (O)
        row[15] = c.line_item_name or c.campaign_name                 # 広告グループ名 (P)
        row[16] = start_dt                                            # 開始時刻 (Q) → 必須
        row[17] = end_dt                                              # 終了時刻 (R)
        row[18] = "PAUSED"                                            # 広告グループの状態 (S) → 一時停止
        # row[19] 広告グループの総予算 (T) → 空欄
        row[20] = _budget_to_number(c.campaign_daily_budget)           # 広告グループの日別予算 (U) → 実際の予算はここ
        row[21] = True                                                # 配信 (V) → 標準配信(True) ※False=集中配信はLINE_ITEM時不可

        # 目標 (Y) - WEBSITE_CONVERSIONS の場合
        if c.campaign_objective == "WEBSITE_CONVERSIONS":
            row[24] = "WEBSITE_CONVERSIONS"                           # 目標 (Y)

        row[25] = _placements_to_excel(c.placements)                  # 広告グループの配信先 (Z)
        row[26] = "PROMOTED_TWEETS"                                   # プロモ商品タイプ (AA)
        row[27] = _api_id_to_editor(c.conversion_tag_id) if c.conversion_tag_id else None  # コンバージョンタグID (AB)
        row[28] = "-"                                                 # お支払い方法 (AC) → "-"

        bid_strategy = BID_STRATEGY_MAP_EXCEL.get(c.bid_strategy, "AUTO_BID")
        row[30] = bid_strategy                                        # 入札戦略 (AE)
        row[31] = float(c.bid_amount) if c.bid_amount else None       # 入札額 (AF) → float
        row[32] = "NONE"                                              # 情報開示のタイプ (AG)

        # === Targeting (37-75) ===
        row[37] = _gender_to_excel(c.target_gender)                   # 性別 (AL)
        row[38] = _age_range_to_excel(c.target_age_ranges)             # 年齢 (AM)
        row[39] = _locations_to_excel(c.target_locations)               # 場所 (AN)
        row[42] = _platforms_to_excel(c.target_platforms)              # プラットフォーム (AQ)
        row[46] = _json_to_semicolon(c.target_languages)              # 言語 (AU)

        # Audiences
        if c.target_audiences:
            row[57] = _json_to_semicolon(c.target_audiences)          # テイラードオーディエンス (BF)

        # 類似オーディエンス拡張
        if c.audience_expansion:
            row[94] = c.audience_expansion                            # 類似オーディエンスの拡張設定 (CQ)

        # Bool fields - X Ads Editor expects bool values
        row[45] = False                                               # Wi-Fiのみ (AT) → False
        row[91] = False                                               # Amplifyプログラムから自動プロモーション (CN) → False

        # === Tweet (76-82) ===
        if c.tweet_ids:
            row[76] = _tweet_ids_to_editor(c.tweet_ids)               # ツイートID (BY)

        return row

    def _build_payment_sheet(
        self,
        ws,
        campaigns: list[SubmissionCampaign],
        project: Optional[Project],
    ):
        """お支払い方法ID シート（元テンプレートと同じヘッダー形式）"""
        headers = ["お支払い方法ID", "お支払い方法名", "ステータス", "開始日", "終了日", "予算"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)

        # ユニークなfunding_instrument_idを列挙
        fi_ids = set()
        for c in campaigns:
            if c.funding_instrument_id:
                fi_ids.add(c.funding_instrument_id)

        for row_idx, fi_id in enumerate(fi_ids, 2):
            editor_id = _api_id_to_editor(fi_id)
            ws.cell(row=row_idx, column=1, value=editor_id)  # お支払い方法ID
            ws.cell(row=row_idx, column=2, value=editor_id)  # お支払い方法名（IDで代用）
            ws.cell(row=row_idx, column=3, value="ACTIVE")   # ステータス

    def _build_conversion_tags_sheet(self, ws, campaigns: list[SubmissionCampaign]):
        """コンバージョンタグID シート（元テンプレートと同じヘッダー形式）"""
        headers = ["コンバージョンタグID", "コンバージョンタグ名"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)

        tag_ids = set()
        for c in campaigns:
            if c.conversion_tag_id:
                tag_ids.add(c.conversion_tag_id)

        for row_idx, tag_id in enumerate(tag_ids, 2):
            editor_id = _api_id_to_editor(tag_id)
            ws.cell(row=row_idx, column=1, value=editor_id)   # コンバージョンタグID
            ws.cell(row=row_idx, column=2, value=editor_id)   # コンバージョンタグ名（IDで代用）

    def _build_audiences_sheet(self, ws):
        """テイラードオーディエンスID シート"""
        headers = ["テイラードオーディエンスID", "名前", "タイプ", "共有可能"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)

    def _build_media_sheet(self, ws):
        """メディアクリエイティブID シート"""
        headers = ["メディアクリエイティブID", "メディアタイプ", "アップロード日"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)
